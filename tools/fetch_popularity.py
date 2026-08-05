# -*- coding: utf-8 -*-
"""
fetch_popularity — omple la columna `pop` del full mestre amb la popularitat de Deezer.

    python tools/fetch_popularity.py             # nomes les que no en tenen
    python tools/fetch_popularity.py --all       # torna a consultar-les totes

`pop` es el camp `rank` de l'API de Deezer (com mes alt, mes escoltada). El fem servir
NOMES per decidir en quin ordre es desbloquegen les cancons per nivell: la mes coneguda
de cada artista, primer. No te cap efecte sobre la puntuacio ni sobre el joc en si.

Diferencia important amb l'any: aqui no ens juguem res si el valor no es exacte. Un
ordre aproximat de popularitat serveix perfectament; una data equivocada, no.
"""
import argparse
import json
import os
import sys
import time
import urllib.request

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import songlib

API = "https://api.deezer.com/track/"
UA = {"User-Agent": "canconet-popularity/1.0"}
COL = "pop"


def track_rank(dzid):
    request = urllib.request.Request(API + str(dzid), headers=UA)
    with urllib.request.urlopen(request, timeout=20) as response:
        data = json.loads(response.read().decode("utf-8", "replace"))
    if data.get("error"):
        return None
    return data.get("rank")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--all", action="store_true", help="refresca totes, no nomes les buides")
    args = parser.parse_args()

    path = songlib.master_path()
    book = openpyxl.load_workbook(path)
    done = missing = skipped = 0

    for lang in songlib.LANGS:
        sheet = book["CAT_" + lang]
        header = [cell.value for cell in sheet[1]]
        if COL not in header:
            sheet.cell(row=1, column=len(header) + 1, value=COL)
            header.append(COL)
        col_pop = header.index(COL) + 1
        col_dz = header.index("dzId") + 1

        for row in range(2, sheet.max_row + 1):
            dzid = songlib.clean_int(sheet.cell(row=row, column=col_dz).value)
            if not dzid:
                continue
            if not args.all and songlib.clean_int(sheet.cell(row=row, column=col_pop).value):
                skipped += 1
                continue
            try:
                rank = track_rank(dzid)
            except Exception:
                rank = None
            if rank:
                sheet.cell(row=row, column=col_pop, value=int(rank))
                done += 1
            else:
                missing += 1
            time.sleep(0.22)
        print("  %-4s fet" % lang)

    book.save(path)
    print("")
    print("popularitat escrita: %d | sense dada: %d | ja en tenien: %d" % (done, missing, skipped))
    print("full mestre:", path)


if __name__ == "__main__":
    main()
