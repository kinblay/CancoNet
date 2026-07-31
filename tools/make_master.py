# -*- coding: utf-8 -*-
"""
make_master — construeix bases/canconet_master.xlsx, l'unic full de calcul
que mana sobre el cataleg.

S'alimenta de:
  * bases/canconet_base_idiomes_r7.xlsx  (any, estil, zona, qualitat, font)
  * index.html                            (id_joc, genere, inicial, st)
  * tools/dzid_proposals.json             (dzId recuperats de Deezer)
  * bases/canconet_roscos*.xlsx           (els tres, fusionats)

Regles:
  * unio sense perdues: si una canco nomes es a un lloc, hi entra igual
  * les cancons del full BROSSA_quarantena queden fora
  * les propostes de Deezer nomes s'apliquen si son ALTA i la fila no es "No apte"
  * id_joc no es reassigna mai (es la clau de la telemetria)

Us:  python tools/make_master.py
"""
import io
import json
import os
import sys
from collections import Counter, defaultdict

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import songlib

R7 = "canconet_base_idiomes_r7.xlsx"
ROSCO_FILES = ["canconet_roscos.xlsx", "canconet_roscos1.xlsx", "canconet_roscos2.xlsx"]

CAT_COLS = ["id_joc", "dzId", "Artista", "Cançó", "Any", "dec", "st", "g", "ai",
            "Zona", "Qualitat", "Source", "Deezer URL", "Notes"]

HEAD_FILL = PatternFill("solid", fgColor="1F2937")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=10)


def load_r7():
    """{lang: {dzId: fila}} i {lang: [files sense dzId]} + dzIds en quarantena."""
    book = openpyxl.load_workbook(songlib.find_base_file(R7), read_only=True, data_only=True)
    by_dz, without = {}, {}
    for lang in songlib.LANGS:
        by_dz[lang], without[lang] = {}, []
        for row in book[lang].iter_rows(min_row=2, values_only=True):
            if not row or not any(cell is not None for cell in row):
                continue
            record = {
                "dzId": songlib.clean_int(row[0]), "Artista": row[1], "Cançó": row[2],
                "Any": songlib.clean_int(row[3]), "Estil": row[4], "Zona": row[5],
                "Qualitat": row[6], "Source": row[7], "Deezer URL": row[8], "Notes": row[9],
            }
            if record["dzId"]:
                by_dz[lang][record["dzId"]] = record
            else:
                without[lang].append(record)
    quarantine = set()
    for row in book["BROSSA_quarantena"].iter_rows(min_row=2, values_only=True):
        if row:
            value = songlib.clean_int(row[0])
            if value:
                quarantine.add(value)
    book.close()
    return by_dz, without, quarantine


def style_map(games, r7_by_dz):
    """Dedueix Estil (slug del full) -> st (categoria del joc) mirant les que ja tenen tots dos."""
    votes = defaultdict(Counter)
    for lang in songlib.LANGS:
        for song in games[lang]:
            excel = r7_by_dz[lang].get(song.get("dzId"))
            if excel and excel.get("Estil") and song.get("st"):
                votes[str(excel["Estil"]).strip()][song["st"]] += 1
    return {slug: counter.most_common(1)[0][0] for slug, counter in votes.items()}


def genre_from_slug(slug):
    """`pop_reggae_catala` -> `Pop reggae català` (nomes per a files noves)."""
    if not slug:
        return ""
    words = str(slug).replace("_", " ").split()
    words = ["català" if word == "catala" else word for word in words]
    return " ".join(words).capitalize()


def build_catalog(games, r7_by_dz, r7_without, quarantine, proposals):
    """Retorna {lang: [dict]} llest per escriure, ja net i complet."""
    slug_to_st = style_map(games, r7_by_dz)
    next_id = max(song["id"] for lang in songlib.LANGS for song in games[lang]) + 1
    report = Counter()
    catalog = {}

    for lang in songlib.LANGS:
        rows, seen = [], set()
        for song in games[lang]:
            dz = song.get("dzId")
            excel = r7_by_dz[lang].get(dz, {})
            if dz in quarantine or str(excel.get("Qualitat") or "").strip() == "No apte":
                report["descartades_quarantena"] += 1
                continue
            seen.add(dz)
            year = song.get("y") or excel.get("Any")
            if not song.get("y") and year:
                report["any_recuperat"] += 1
            estil = excel.get("Estil")
            st_value = song.get("st") or slug_to_st.get(str(estil).strip(), "")
            if not song.get("st") and st_value:
                report["st_deduit"] += 1
            rows.append({
                "id_joc": song["id"], "dzId": dz,
                "Artista": song.get("a") or excel.get("Artista"),
                "Cançó": song.get("t") or excel.get("Cançó"),
                "Any": year, "dec": songlib.decade_label(year),
                "st": st_value, "g": song.get("g") or genre_from_slug(estil),
                "ai": song.get("ai") or (str(song.get("a") or " ")[0].upper()),
                "Zona": excel.get("Zona"), "Qualitat": excel.get("Qualitat"),
                "Source": excel.get("Source"), "Deezer URL": excel.get("Deezer URL"),
                "Notes": excel.get("Notes"),
            })

        # files que nomes son a l'Excel (tenen dzId pero no son al joc)
        for dz, excel in r7_by_dz[lang].items():
            if dz in seen or dz in quarantine:
                continue
            if str(excel.get("Qualitat") or "").strip() == "No apte":
                continue
            report["afegides_des_excel"] += 1
            rows.append({
                "id_joc": next_id, "dzId": dz, "Artista": excel.get("Artista"),
                "Cançó": excel.get("Cançó"), "Any": excel.get("Any"),
                "dec": songlib.decade_label(excel.get("Any")),
                "st": slug_to_st.get(str(excel.get("Estil")).strip(), ""),
                "g": genre_from_slug(excel.get("Estil")),
                "ai": str(excel.get("Artista") or " ")[0].upper(),
                "Zona": excel.get("Zona"), "Qualitat": excel.get("Qualitat"),
                "Source": excel.get("Source"), "Deezer URL": excel.get("Deezer URL"),
                "Notes": excel.get("Notes"),
            })
            next_id += 1
        catalog[lang] = rows

    # dzId recuperats de Deezer: nomes els d'alta confianca i mai els "No apte"
    applied, pending = [], []
    known = {row["dzId"] for lang in songlib.LANGS for row in catalog[lang]}
    for item in proposals:
        source_row = next((rec for rec in r7_without["CA"]
                           if songlib.norm(rec["Artista"]) == songlib.norm(item["artista"])
                           and songlib.norm(rec["Cançó"]) == songlib.norm(item["canco"])), {})
        blocked = str(source_row.get("Qualitat") or "").strip() == "No apte"
        if item["estat"] == "ALTA" and item["dzId"] and item["dzId"] not in known and not blocked:
            catalog["CA"].append({
                "id_joc": next_id, "dzId": item["dzId"], "Artista": item["artista"],
                "Cançó": item["canco"], "Any": songlib.clean_int(item.get("any")),
                "dec": songlib.decade_label(item.get("any")),
                "st": slug_to_st.get(str(item.get("estil")).strip(), ""),
                "g": genre_from_slug(item.get("estil")),
                "ai": str(item["artista"] or " ")[0].upper(),
                "Zona": source_row.get("Zona") or "cat",
                "Qualitat": "Base ✓ (dzId recuperat)", "Source": "deezer_api_%d%%" % item["confianca"],
                "Deezer URL": "https://www.deezer.com/track/%s" % item["dzId"],
                "Notes": "dzId trobat automaticament (%s - %s)" % (item.get("deezer_artista"), item.get("deezer_titol")),
            })
            known.add(item["dzId"])
            applied.append(item)
            next_id += 1
        else:
            pending.append(item)
    report["dzid_aplicats"] = len(applied)
    report["dzid_pendents"] = len(pending)
    return catalog, pending, report


def read_rosco_sheets():
    """Fusiona els tres fitxers de roscos. El primer que porta un full, mana."""
    merged = {}
    for filename in ROSCO_FILES:
        path = songlib.find_base_file(filename)
        if not path:
            continue
        book = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for name in book.sheetnames:
            if name in ("INDEX", "PLANTILLA") or name in merged:
                continue
            rows = [row for row in book[name].iter_rows(values_only=True)
                    if row and any(cell is not None for cell in row)]
            if rows:
                merged["ROSCO_" + name] = rows
        book.close()
    return merged


def write_sheet(book, title, header, rows, widths=None):
    sheet = book.create_sheet(title)
    sheet.append(header)
    for cell in sheet[1]:
        cell.fill, cell.font = HEAD_FILL, HEAD_FONT
        cell.alignment = Alignment(vertical="center")
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    for index, width in enumerate(widths or [], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    return sheet


def main():
    games = songlib.load_songs()
    r7_by_dz, r7_without, quarantine = load_r7()
    proposal_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "dzid_proposals.json")
    proposals = []
    if os.path.exists(proposal_file):
        with io.open(proposal_file, encoding="utf-8") as handle:
            proposals = json.load(handle)

    catalog, pending, report = build_catalog(games, r7_by_dz, r7_without, quarantine, proposals)

    book = openpyxl.Workbook()
    book.remove(book.active)

    index_rows = [
        ["CAT_CA / CAT_ES / CAT_EU / CAT_INT", "Cataleg de cancons. Genera els arrays SONGS_* d'index.html amb: python tools/build_songs.py"],
        ["ROSCO_*", "Preguntes dels roscos. Van a la taula rosco_questions de Supabase."],
        ["PENDENTS_DZID", "Cancons validades sense dzId segur. Revisa la proposta i, si es bona, copia el dzId a la fila del cataleg."],
        ["", ""],
        ["Regla d'or", "id_joc no es canvia mai: es el que guarda la telemetria (game_rounds.song_id)."],
        ["dec", "Es calcula sol a partir de l'Any. Nomes es text de pista dins del joc."],
        ["Camps retirats", "seg i ytId ja no s'usen (el joc reprodueix previews de 30 s amb finestres fixes per nivell)."],
    ]
    write_sheet(book, "_INDEX", ["Full", "Per a que serveix"], index_rows, [34, 104])

    for lang in songlib.LANGS:
        rows = sorted(catalog[lang], key=lambda item: item["id_joc"])
        write_sheet(book, "CAT_" + lang, CAT_COLS,
                    [[row.get(col) for col in CAT_COLS] for row in rows],
                    [8, 13, 26, 34, 7, 11, 10, 22, 5, 7, 22, 16, 46, 40])

    write_sheet(book, "PENDENTS_DZID",
                ["Artista", "Cançó", "Any", "Estil", "Estat", "Confiança %",
                 "dzId proposat", "Deezer diu (artista)", "Deezer diu (títol)", "Validat?"],
                [[item["artista"], item["canco"], item["any"], item["estil"], item["estat"],
                  item["confianca"], item["dzId"], item["deezer_artista"], item["deezer_titol"], ""]
                 for item in pending],
                [24, 30, 7, 22, 10, 12, 14, 24, 30, 10])

    for name, rows in read_rosco_sheets().items():
        write_sheet(book, name[:31], list(rows[0]), [list(row) for row in rows[1:]],
                    [8, 12, 60, 26, 26, 12, 30])

    book.save(songlib.master_path())

    print("=== MASTER CREAT ===")
    print(songlib.master_path())
    print("")
    for lang in songlib.LANGS:
        print("  CAT_%-4s %4d cancons" % (lang, len(catalog[lang])))
    print("  TOTAL    %4d" % sum(len(catalog[lang]) for lang in songlib.LANGS))
    print("")
    print("  anys recuperats de l'Excel : %d" % report["any_recuperat"])
    print("  st deduits del slug d'estil: %d" % report["st_deduit"])
    print("  descartades (quarantena)   : %d" % report["descartades_quarantena"])
    print("  afegides des de l'Excel    : %d" % report["afegides_des_excel"])
    print("  dzId recuperats aplicats   : %d" % report["dzid_aplicats"])
    print("  dzId encara pendents       : %d" % report["dzid_pendents"])


if __name__ == "__main__":
    main()
