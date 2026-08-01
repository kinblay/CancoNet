# -*- coding: utf-8 -*-
"""
build_songs — porta el cataleg del full mestre al joc.

    python tools/build_songs.py            # valida, escriu index.html i data/*.csv
    python tools/build_songs.py --check    # nomes valida i informa (no escriu res)

Llegeix bases/canconet_master.xlsx (fulls CAT_CA, CAT_ES, CAT_EU, CAT_INT) i
regenera NOMES els arrays SONGS_* d'index.html. La resta del fitxer no es toca.

Comprovacions que atura la publicacio (errors):
  * id_joc repetit (trencaria la telemetria i el cercador de suggeriments)
  * dzId repetit dins del mateix idioma
  * files sense dzId o sense titol/artista
  * cancons que tornen a apareixer venint de la quarantena
"""
import argparse
import csv
import io
import os
import sys
from collections import Counter

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import songlib

R7 = "canconet_base_idiomes_r7.xlsx"
CSV_COLS = ["id", "dzId", "t", "a", "y", "g", "ai", "dec", "st"]


def quarantine_ids():
    """dzIds que algu va apartar per dolents; no han de tornar mai."""
    path = songlib.find_base_file(R7)
    if not path:
        return set()
    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = set()
    if "BROSSA_quarantena" in book.sheetnames:
        for row in book["BROSSA_quarantena"].iter_rows(min_row=2, values_only=True):
            if row:
                value = songlib.clean_int(row[0])
                if value:
                    out.add(value)
    book.close()
    return out


def read_master():
    book = openpyxl.load_workbook(songlib.master_path(), read_only=True, data_only=True)
    catalog = {}
    for lang in songlib.LANGS:
        sheet = book["CAT_" + lang]
        header = [str(cell) for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        rows = []
        for values in sheet.iter_rows(min_row=2, values_only=True):
            if not values or not any(cell is not None for cell in values):
                continue
            rows.append(dict(zip(header, values)))
        catalog[lang] = rows
    book.close()
    return catalog


def to_song(row):
    """Fila del full -> objecte tal com el vol el joc."""
    song = {
        "id": songlib.clean_int(row.get("id_joc")),
        "dzId": songlib.clean_int(row.get("dzId")),
        "t": (row.get("Cançó") or "").strip(),
        "a": (row.get("Artista") or "").strip(),
        "y": songlib.clean_int(row.get("Any")),
        "g": (row.get("g") or "").strip(),
        "ai": (row.get("ai") or "").strip(),
        "dec": (row.get("dec") or "").strip(),
        "st": (row.get("st") or "").strip(),
    }
    # la decada surt sempre de l'any, per no tenir dues etiquetes per a la mateixa cosa
    song["dec"] = songlib.decade_label(song["y"]) if song["y"] else ""
    return song


def validate(catalog, quarantine):
    errors, warnings = [], []
    all_ids = Counter()
    for lang in songlib.LANGS:
        dz_seen = Counter()
        for row in catalog[lang]:
            song = to_song(row)
            label = "%s/%s - %s" % (lang, song["a"][:20], song["t"][:28])
            if not song["id"]:
                errors.append("sense id_joc: " + label)
            else:
                all_ids[song["id"]] += 1
            if not song["dzId"]:
                errors.append("sense dzId: " + label)
            else:
                dz_seen[song["dzId"]] += 1
                if song["dzId"] in quarantine:
                    errors.append("torna de la quarantena: " + label)
            if not song["t"] or not song["a"]:
                errors.append("sense titol o artista: " + label)
            if not song["y"]:
                warnings.append("sense any (no se li preguntara l'any): " + label)
            if not song["st"]:
                warnings.append("sense estil (no surt al filtre del mode lliure): " + label)
        for dz, count in dz_seen.items():
            if count > 1:
                errors.append("dzId repetit a %s: %s (%d cops)" % (lang, dz, count))
    for song_id, count in all_ids.items():
        if count > 1:
            errors.append("id_joc repetit: %s (%d cops)" % (song_id, count))
    return errors, warnings


def write_csvs(catalog):
    folder = os.path.join(songlib.repo_root(), "data")
    if not os.path.isdir(folder):
        os.makedirs(folder)
    for lang in songlib.LANGS:
        path = os.path.join(folder, "songs_%s.csv" % lang.lower())
        with io.open(path, "w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=CSV_COLS)
            writer.writeheader()
            for row in sorted(catalog[lang], key=lambda item: songlib.clean_int(item.get("id_joc")) or 0):
                writer.writerow(to_song(row))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--check", action="store_true", help="nomes validar, no escriure")
    args = parser.parse_args()

    catalog = read_master()
    errors, warnings = validate(catalog, quarantine_ids())

    print("=== VALIDACIO ===")
    for lang in songlib.LANGS:
        rows = catalog[lang]
        with_year = sum(1 for row in rows if songlib.clean_int(row.get("Any")))
        with_style = sum(1 for row in rows if (row.get("st") or "").strip())
        print("  CAT_%-4s %4d cancons | amb any %4d | amb estil %4d" % (lang, len(rows), with_year, with_style))
    print("  TOTAL    %4d" % sum(len(catalog[lang]) for lang in songlib.LANGS))
    print("")
    print("  errors  : %d" % len(errors))
    for item in errors[:15]:
        print("     ! " + item)
    if len(errors) > 15:
        print("     ... i %d mes" % (len(errors) - 15))
    print("  avisos  : %d  (no aturen la publicacio)" % len(warnings))
    for item in warnings[:5]:
        print("     - " + item)
    if len(warnings) > 5:
        print("     ... i %d mes" % (len(warnings) - 5))

    if errors:
        print("")
        print("ATURAT: arregla els errors al full mestre i torna-ho a executar.")
        return 1
    if args.check:
        print("")
        print("--check: tot correcte, no s'ha escrit res.")
        return 0

    songs_by_lang = {}
    for lang in songlib.LANGS:
        rows = sorted(catalog[lang], key=lambda item: songlib.clean_int(item.get("id_joc")) or 0)
        songs_by_lang[lang] = [to_song(row) for row in rows]

    source = songlib.read_index()
    before = len(source)
    songlib.write_index(songlib.replace_arrays(source, songs_by_lang))
    write_csvs(catalog)
    after = len(songlib.read_index())
    print("")
    print("=== ESCRIT ===")
    print("  index.html: %d -> %d bytes (%+d KB)" % (before, after, (after - before) / 1024))
    print("  data/songs_*.csv actualitzats")
    print("")
    print("  RECORDA: puja la versio de la cache a sw.js o els jugadors no ho veuran.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
